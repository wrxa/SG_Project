# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from util.get_all_path import GetPath


class Write_excel(object):

    def __init__(self, filename, sheetindex):
        self.filename = filename
        self.wb = load_workbook(self.filename)
        self.wb.active = sheetindex
        self.ws = self.wb.active

    def write(self, coordDir, targetpath):
        # eg: coord:A1
        for coordkey in coordDir:
            self.ws[coordkey].value = coordDir[coordkey]
        self.wb.save(targetpath)


def createEconomicExcle(planId, userId):
    from app.ccpp.models.ccpp_ccpp_economicModel import Ccpp_ccpp_economic
    from app.ccpp import gl
    economicccpp = Ccpp_ccpp_economic.search_economic(planId)
    full_year = economicccpp.full_investment_yield_year
    depreciation_years = economicccpp.depreciation_amortization_years
    wr = Write_excel(GetPath.getExcleCcppSource(full_year, depreciation_years), 0)
    coordDir = {}
    coordDir['M4'] = gl.format_value_forexcle(economicccpp.natural_gas_price)
    coordDir['M5'] = gl.format_value_forexcle(economicccpp.coal_price)
    coordDir['M6'] = gl.format_value_forexcle(economicccpp.water_price)
    coordDir['M7'] = gl.format_value_forexcle(economicccpp.electricity_price)
    coordDir['M9'] = gl.format_value_forexcle(economicccpp.energy_priceother)
    coordDir['R4'] = gl.format_value_forexcle(economicccpp.energy_supply_price)
    coordDir['R5'] = gl.format_value_forexcle(economicccpp.energy_supply_water_price)
    coordDir['R6'] = gl.format_value_forexcle(economicccpp.energy_supply_for_heating)
    coordDir['R7'] = gl.format_value_forexcle(economicccpp.energy_supply_refrigeration_price)
    coordDir['R8'] = gl.format_value_forexcle(economicccpp.energy_supply_steam_price)
    coordDir['R9'] = gl.format_value_forexcle(economicccpp.other_energy_supply)

    coordDir['C19'] = gl.format_value_forexcle(economicccpp.value_added_tax_rate_power_supply / 100 if economicccpp.value_added_tax_rate_power_supply is not None else None)
    coordDir['C20'] = gl.format_value_forexcle(economicccpp.value_added_tax_rate_heating_water / 100 if economicccpp.value_added_tax_rate_heating_water is not None else None)
    coordDir['C21'] = gl.format_value_forexcle(economicccpp.heating_income_vat / 100 if economicccpp.heating_income_vat is not None else None)
    coordDir['C22'] = gl.format_value_forexcle(economicccpp.vat_for_cooling_income / 100 if economicccpp.vat_for_cooling_income is not None else None)
    coordDir['C23'] = gl.format_value_forexcle(economicccpp.vat_for_steam_income / 100 if economicccpp.vat_for_steam_income is not None else None)
    coordDir['C30'] = gl.format_value_forexcle(economicccpp.gas_cost_value_added_tax_rate / 100 if economicccpp.gas_cost_value_added_tax_rate is not None else None)
    coordDir['C31'] = gl.format_value_forexcle(economicccpp.coal_cost_value_added_tax_rate / 100 if economicccpp.coal_cost_value_added_tax_rate is not None else None)
    coordDir['C32'] = gl.format_value_forexcle(economicccpp.power_cost_value_added_tax_rate / 100 if economicccpp.power_cost_value_added_tax_rate is not None else None)
    coordDir['C33'] = gl.format_value_forexcle(economicccpp.water_consumption_cost_value_added_tax_rate / 100 if economicccpp.water_consumption_cost_value_added_tax_rate is not None else None)
    coordDir['C38'] = gl.format_value_forexcle(economicccpp.depreciation_amortization_years)
    coordDir['C43'] = gl.format_value_forexcle(economicccpp.post_tax_profit_vat / 100 if economicccpp.post_tax_profit_vat is not None else None)
    coordDir['C47'] = gl.format_value_forexcle(economicccpp.percentage_capital_total_investment / 100 if economicccpp.percentage_capital_total_investment is not None else None)
    coordDir['C50'] = gl.format_value_forexcle(economicccpp.interest_rate_bank_interest / 100 if economicccpp.interest_rate_bank_interest is not None else None)
    coordDir['D45'] = gl.format_value_forexcle(economicccpp.purchase_cost)
    coordDir['D46'] = gl.format_value_forexcle(economicccpp.new_investment)
    zmDir = {'D': '1', 'E': '2', 'F': '3', 'G': '4', 'H': '5', 'I': '6', 'J': '7', 'K': '8', 'L': '9', 'M': '10',
             'N': '11', 'O': '12', 'P': '13', 'Q': '14', 'R': '15', 'S': '16', 'T': '17', 'U': '18', 'V': '19', 'W': '20'}
    zddir = {'13': 'power_supply_capacity', '14': 'energy_supply_for_heating', '15': 'energy_supply_heating',
             '16': 'energy_supply_for_cooling', '17': 'vapour_production', '18': 'income_other',
             '24': 'income_otherother', '26': 'gas_consumption', '27': 'coal_consumption', '28': 'power_consumption',
             '29': 'water_consumption', '34': 'material_cost', '35': 'maintenance_cost', '36': 'artificial_cost',
             '37': 'otherincluding_rent_capacity', '51': 'repayment_plan'}
    for zmkey in zmDir:
        for zdkey in zddir:
            coordDir['' + zmkey + '' + zdkey + ''] = gl.format_value_forexcle(getattr(economicccpp, zddir[zdkey] + '_' + zmDir[zmkey]))
    wr.write(coordDir, GetPath.getExcleCcppResult(planId, userId))


def main():
    wr = Write_excel("ccppbaseexcle/economicccpp.xlsx", 0)
    wr.write({'M4': None}, str(1) + "-economicccpp.xlsx")


if __name__ == '__main__':
    main()
   